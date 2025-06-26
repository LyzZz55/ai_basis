# utils.py
# for alpha ver 1.15.0+ [Main] Agent 2

import json
from typing import Dict, Any



def output(color: str, message: str, f=None, std_flag=None):
    """一个简单的打印函数，用于模拟给定的 output 函数。"""
    color_map = {
        "GREEN": "\033[92m",
        "RED": "\033[91m",
        "END": "\033[0m",
    }
    start_color = color_map.get(color, "")
    end_color = color_map.get("END", "")
    print(f"{start_color}{message}{end_color}")


def load_file_config(json_path: str) -> Dict[str, Any]:
    """从 JSON 文件加载文件配置"""
    print(f"正在从 '{json_path}' 加载配置文件...")
    with open(json_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    return config

import pandas as pd
from PIL import Image
def load_files_from_config(config: Dict[str, Any]) -> Dict[str, Any]:
    """
    根据配置加载所有文件内容。
    """
    file_contents = {}
    print("开始根据配置加载输入文件...")

    for file_info in config.get("files", []):
        file_path = file_info["path"]
        comment = file_info.get("comment", "")
        file_type = file_info.get("type", "")
        
        try:
            if file_type == 'text':
                # --- 核心修改：使用标准 Python 读取文件 ---
                # 直接以文本模式（'r'）和 utf-8 编码打开文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    # 读取文件的全部内容为字符串
                    text_content = f.read()

                # 直接将读取到的字符串内容存入字典
                file_contents[comment] = {
                    "content": text_content,  # 直接使用读取到的字符串
                    "path": file_path,
                    "type": file_type
                }
            elif file_type == 'excel':
                 # 读取所有sheet为dict
                excel_data = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
                # 转为dict（sheet名: 行数据list[dict]）
                excel_dict = {sheet: df.to_dict(orient="records") for sheet, df in excel_data.items()}
                file_contents[file_path] = {
                    "content": excel_dict,
                    "comment": comment,
                    "type": file_type
                }
            elif file_type == 'img':
                img = Image.open(file_path)
                file_contents[file_path] = {
                    "content": img.copy(),  # 返回PIL对象副本
                    "comment": comment,
                    "type": file_type
                }
                img.close()
            else:
                output("RED", f"✘ 文件类型不明: {file_path}")
                return {}
            output("GREEN", f"✔ 已成功加载文件: '{file_path}' ({comment})")
        except FileNotFoundError:
            output("RED", f"✘ 文件未找到: {file_path}")
            return {}
        except UnicodeDecodeError:
            output("RED", f"✘ 文件 '{file_path}' 不是有效的 UTF-8 编码，无法解析。")
            return {}
        except Exception as e:
            output("RED", f"✘ 加载文件 {file_path} 时发生未知错误: {e}")
            return {}
    print("✅ 所有输入文件加载完毕。")
    return file_contents

################



import os

def save_to_file(content: str, *path_parts) -> None:
    """将字符串保存到指定多层路径文件，自动创建文件夹"""
    file_path = os.path.join(*path_parts)
    dir_path = os.path.dirname(file_path)
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)
    try:
        with open(file_path, 'a', encoding='utf-8') as file:
            file.write(content)
        print(f"内容已追加到文件: {file_path}")
    except Exception as e:
        print(f"写入文件时出错: {e}")

from pathlib import Path
import openpyxl
def xlsx_to_json(excel_file_path, output_file=None, sheet_name=None):
    """
    将Excel文件转换为JSON格式
    
    Args:
    - excel_file_path: Excel文件路径
    - output_file: 输出JSON文件路径，默认为None(直接返回JSON字符串)
    - sheet_name: 要处理的表名，默认为None(使用第一个表)
    """
    try:
        # 检查文件是否存在
        if not Path(excel_file_path).exists():
            raise FileNotFoundError(f"文件不存在: {excel_file_path}")
        
        # 打开工作簿
        wb = openpyxl.load_workbook(excel_file_path, read_only=True)
        
        # 获取工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            ws = wb[sheet_name]
        else:
            ws = wb.active  # 使用第一个工作表
        
        # 获取表头(第一行)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        
        # 检查是否有表头
        if not any(headers):
            raise ValueError("Excel文件不包含表头")
        
        # 读取数据行
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not any(row):
                continue
                
            # 创建字典
            row_dict = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = value
            
            data.append(row_dict)
        
        # 关闭工作簿
        wb.close()
        
        # 转换为JSON
        json_data = json.dumps(data, ensure_ascii=False, indent=2)
        
        # 输出到文件或返回
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(json_data)
            print(f"JSON数据已保存到: {output_file}")
        else:
            return json_data
    
    except Exception as e:
        print(f"发生错误: {str(e)}")
        return None